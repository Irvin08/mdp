##Changes needed:
##Check for most recent leads passdown
##add mayan test percent checker
##expand on chamber class
##refactor to use chamber class when possibe
##Look into SAP connection
##Find better way to add last test run, currently takes about 1-1.5 sec
##Keep track of where to update changesot system num/ctv etc, maybe master list
##look into updating based on gantt

import pandas as pd
from datetime import datetime, timedelta
from PIL import Image, ImageDraw
import tkinter as tk
from tkinter import *
from PIL import ImageTk
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import webbrowser
import win32com.client as win32
import os.path
from os import path
import glob
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
import fnmatch
import time


class Chamber:
    def __init__ (self, system, chamberPO, ctvPO, gpPO, chType, portIdx):
        self.system = system
        self.chPO = chamberPO
        self.ctvPO = ctvPO
        self.gpPO = gpPO
        self.chType = chType
        self.portIdx = portIdx
        self.gpQNs = []
        self.ctvQNs = []
        self.QNs = []
        self.InspLots = []
        self.ESWs = []
        self.allQNs = [self.QNs, self.ctvQNs]

    def __str__(self):
        return 'System #:{}, PO #: {}'.format(self.system,self.chPO)

    def print(self):
        print(self.system + " " + self.chPO + " " + self.chType)

    def numOpenQNs(self, qnIdx):
        openQNs = 0
        for q in self.allQNs[qnIdx]:
            if q.isOpen:
                openQNs = openQNs + 1
        return openQNs

    def openQNs(self, qnIdx):
        qnList = []
        for q in self.allQNs[qnIdx]:
            if q.isOpen:
                qnList.append(q)
        return qnList
    
    def QNStatus(self, onlyOpen, status, qnIdx):
        if not self.allQNs[qnIdx]:
            return "There are no QN's\n"
        allClosed = True
        if onlyOpen:
            for q in self.allQNs[qnIdx]:
                if q.isOpen:
                    status = status + q.print()
                    allClosed = False
        else:
            for q in self.allQNs[qnIdx]:
                if q.isOpen:
                    allClosed = False
                status = status + q.print()
        if onlyOpen and allClosed:
            status = "There are no open QN's\n"
        return status

class Port:
    def __init__ (self, portName, occupied, chamber, chamberPO, chamberType):
        self.portName = portName
        self.occupied = occupied
        self.chamber = chamber
        self.chamberPO = chamberPO
        self.chamberType = chamberType

    def info(self):
        return self.portName + '|' + self.chamber + '|' +self.chamberPO + '|' + self.chamberType + '\n'
    
    def print(self):
        print(self.portName + ' ' + self.chamber + ' ' +self.chamberPO)


class QN:
    def __init__ (self, QNNum, Type, Desc, isOpen, partNum = None):
        self.QNNum = QNNum
        self.Type = Type
        self.Desc = Desc
        self.isOpen = isOpen
        self.partNum = partNum
        self.dateClosed = None
        self.creator = None
        
    def isOpen(self):
        return self.isOpen

    def print(self):
        return (self.Type + " " + self.QNNum + " - " + self.Desc.upper() + ("\n\n" if self.isOpen else "CLOSED\n\n"))


class HyperlinkManager:

    def __init__(self, text):

        self.text = text

        self.text.tag_config("hyper", foreground="blue", underline=1)

        self.text.tag_bind("hyper", "<Enter>", self._enter)
        self.text.tag_bind("hyper", "<Leave>", self._leave)
        self.text.tag_bind("hyper", "<Button-1>", self._click)

        self.reset()

    def reset(self):
        self.links = {}

    def add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hyper-%d" % len(self.links)
        self.links[tag] = action
        return "hyper", tag

    def _enter(self, event):
        self.text.config(cursor="hand2")

    def _leave(self, event):
        self.text.config(cursor="")

    def _click(self, event):
        for tag in self.text.tag_names(CURRENT):
            if tag[:6] == "hyper-":
                self.links[tag]()
                return


def getLatestFile(path):
    list_of_files = glob.glob(path)
    if not list_of_files:
        return None
    latest_file = max(list_of_files, key=os.path.getctime)
    _, filename = os.path.split(latest_file)
    return filename


def printStatus(system):
    print(system)
    rows = rackStatusSheet.iter_rows(rackStatusSheet.max_row - 100, rackStatusSheet.max_row)
    status = ""
    found = False
    for row in rows:
        if row[0].value == system:
            found = True
            for x in range(8,12):
                if "NA" in str(row[x].value):
                    print("There is no EQRK #" + str(x - 7) + "\n")
                    status = status + "There is no EQRK #" + str(x - 7) + "\n --------------------------------------\n"
                elif row[x].value == None:
                    print("EQRK #" + str(x - 7) + " is not ready, needed for:")
                    status = status + "EQRK #" + str(x - 7) + " is not ready, needed for:"
                    x = str(row[x].comment)
                    y = x.find(':')
                    y = x.find(':', y + 1)
                    y = y + 1
                    z = x.find('by')
                    print(x[y:z-1] + "\n")
                    status = status + x[y:z-1] + "\n --------------------------------------\n"
                else:
                    print("EQRK #" + str(x - 7) + " is complete for:")
                    status = status + "EQRK #" + str(x - 7) + " is complete for:"
                    #May want to get completed date: row[x].value
                    x = str(row[x].comment)
                    y = x.find(':')
                    y = x.find(':', y + 1)
                    y = y + 1
                    z = x.find('by')
                    print(x[y:z-1] + "\n")
                    status = status + x[y:z-1] + "\n --------------------------------------\n"
            try:
                print("Comments: " + row[12].value)
                status = status + "Comments:\n" + row[12].value + "\n"
                ##ADD CHECK IF NEXT ROW IS EMPTY, IF SO NEED TO GET NEXT COMMENT 
            except:
                pass
    if not found:
        status = "EQRK not started"
        print(status)
    window =tk.Toplevel(root)
    window.geometry('+%d+%d' % (690, 100))
    rack_label = tk.Label(window, text = status)
    rack_label.pack()
    quit_button_rack = tk.Button(window, text = "quit", command = window.destroy)
    quit_button_rack.pack(side = "left")
    window.focus_set()                                                        
    window.grab_set()
    
#maybe add build %
def getPOs(chambers):
    global dfCrossover
    global df
    global dfManual
    manualPOs = load_workbook(manualPOsFile)
    ws = manualPOs.active
    rows = list(ws.rows)
    global manualRows
    for x in range(0, 60):
        skip = False
        done = False
        y = df.at[x, 'System #']
        y = y.replace(" ", "")
        if y == "EMPTY":
            chambers.append(Chamber("XXXXX-X", "XXXXX", "XXXXX", "XXXXX", "XXXXX", x))
            skip = True
        if not skip:
            z = y.find('-')
            y = y[0 : z + 2]
            print(y)
            if "D01" in y:
                try:
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '] == y]
                    chambers.append(Chamber(y, r.values[0,1], "XXXXX", "XXXXX", r.values[0,2], x))
                except IndexError:
                    y = y.replace("D01", "")
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '] == y]
                    chambers.append(Chamber(y, r.values[0,1], "XXXXX", "XXXXX", r.values[0,2], x))
            else:
                try:
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '] == y]
                    po = r.values[0,1]
                    chtype = r.values[0,2]
                    found = True
                except IndexError:
                    pass
            r = rows[x + 1]
            if r[1].value == y:
                init = [y, po, None, None, chtype, x]
                z = [cell.value for cell in r]
                for i in range(1,6):
                    if z[i]:
                        init[i-1] = z[i] 
                chambers.append(Chamber(*init))
                done = True
            if not done:
                chambers.append(Chamber(y, po, "XXXXX", "XXXXX", chtype, x)) if found else chambers.append(Chamber(y, "XXXXX", "XXXXX", "XXXXX", "XXXXXX", x))
                

def getPriorityColors(file, cells):
    wb = load_workbook(file, data_only = True)
    sh = wb['Lead Passdown']
    for x in range (61):
        color_in_hex = sh["A" + str(x+1)].fill.start_color.index
        if color_in_hex != 0:
            try:
                rgb = tuple(int(color_in_hex[i:i+2], 16) for i in ( 2, 4, 6))
                cells.append(rgb)
            except TypeError:
                rgb = (255, 255, 255)
                cells.append(rgb)
        else:
            cells.append((255, 255, 255))
    del cells[0]


def findChamberLocations(bay):
    global chamber_locations
    global df
    for x in range (6):
        if "EMPTY" in df.at[x + (6 * (bay - 1)),'System #']:
            chamber_locations[x] = 0
        else:
            chamber_locations[x] = 1

    return chamber_locations


def create_buttons(root, chamber_image, chamber_locations, active_buttons):
    if chamber_locations[0] == 1:
        portA_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(0)), anchor = "w")
        portA_button_window = canvas.create_window(890,75, anchor= "nw", window = portA_button)
        active_buttons.append(portA_button)

    if chamber_locations[1] == 1:
        portB_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(1)), anchor = "w")
        portB_button_window = canvas.create_window(890,340, anchor= "nw", window = portB_button)
        active_buttons.append(portB_button)

    if chamber_locations[2] == 1:
        portC_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(2)), anchor = "w")
        portC_button_window = canvas.create_window(890,605, anchor= "nw", window = portC_button)
        active_buttons.append(portC_button)

    if chamber_locations[5] == 1:
        portF_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(5)), anchor = "w")
        portF_button_window = canvas.create_window(100,75, anchor= "nw", window = portF_button)
        active_buttons.append(portF_button)

    if chamber_locations[4] == 1:
        portE_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(4)), anchor = "w")
        portE_button_window = canvas.create_window(100,340, anchor= "nw", window = portE_button)
        active_buttons.append(portE_button)

    if chamber_locations[3] == 1:
        portD_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(3)), anchor = "w")
        portD_button_window = canvas.create_window(100,605, anchor= "nw", window = portD_button)
        active_buttons.append(portD_button)


def create_window_Generic(x):
    window =tk.Toplevel(root)
    window.geometry('+%d+%d' % (690, 100))
    chamber = chambers[x + (6 * (bay_num - 1))]
    chamberPO = chamber.chPO
    ctvPO = chamber.ctvPO
    gpPO = chamber.gpPO
    if cells[x + (6 * (bay_num - 1))] == (0, 0, 0):
        cells[x + (6 * (bay_num - 1))] = (255, 255, 255)
    PriorityLabel = tk.Label(window, background = ("#%02x%02x%02x" % cells[x + (6 * (bay_num - 1))]))
    PriorityLabel.grid(row = 0, column = 0, columnspan = 4, sticky = "nsew")#pack()
    portNumber = x + (6 * (bay_num - 1))
    Portlabel = tk.Label(window, text = "Port: " + bay_num_str + ports[x])
    Portlabel.grid(row = 1, column = 1, columnspan = 2, sticky = "nsew")#pack()
    system = df.at[x + (6 * (bay_num - 1)),'System #']
    systemEntry = tk.Text(window, height = 1, width = 25, borderwidth = 0)
    systemEntry.insert(1.0, system)
    systemEntry.tag_configure("center", justify='center')
    systemEntry.tag_add("center", "1.0", "end")
    systemEntry.configure(state="disabled")
    systemEntry.configure(bg=window.cget('bg'), relief="flat")
    systemEntry.grid(row = 2, column = 1, columnspan = 2, sticky = "nsew")
    #Systemlabel = tk.Label(window, text = "System #: " + system)
    #Systemlabel.grid(row = 2, column = 1, columnspan = 2, sticky = "nsew")
    system = system[0 : system.find("-")]
    #
    chTypeText = chamber.chType if chamber.chType is not None else "No chamber type found"
    chamberTypeLabel = tk.Label(window, text = "CH Type: " + chTypeText)
    chamberTypeLabel.grid(row = 3, column = 1, columnspan = 2, sticky = "nsew")#pack()
    POEntry = tk.Text(window, height = 1, width = 25, borderwidth = 0)
    if chamberPO is None:
        POEntry.insert(1.0, "PO is not updated, please update PO#.xlsx")
    else:
        POEntry.insert(1.0, "Chamber PO: " + chamberPO)
    POEntry.tag_configure("center", justify='center')
    POEntry.tag_add("center", "1.0", "end")
    POEntry.configure(state="disabled")
    POEntry.configure(bg=window.cget('bg'), relief="flat")
    POEntry.grid(row = 4, column = 1, columnspan = 1, sticky = "nsew")
    ctvPOEntry = tk.Text(window, height = 1, width = 25, borderwidth = 0)
    ctvText = "CTV PO is not updated" if (ctvPO == "XXXXX" or ctvPO == None) else "CTV PO: " + chamber.ctvPO
    ctvPOEntry.insert(1.0, ctvText)
    ctvPOEntry.tag_configure("center", justify = 'center')
    ctvPOEntry.tag_add("center", "1.0", "end")
    ctvPOEntry.configure(state="disabled")
    ctvPOEntry.configure(bg=window.cget('bg'), relief="flat")
    ctvPOEntry.grid(row = 4, column = 2, sticky = "nsew")
    gpPOEntry = tk.Text(window, height = 1, width = 25, borderwidth = 0)
    gpText = "No Gas Panel PO available" if (gpPO == "XXXXX" or gpPO == None) else "GP PO: " + chamber.gpPO
    gpPOEntry.insert(1.0, gpText)
    gpPOEntry.tag_configure("center", justify = 'center')
    gpPOEntry.tag_add("center", "1.0", "end")
    gpPOEntry.configure(state="disabled")
    gpPOEntry.configure(bg=window.cget('bg'), relief="flat")
    gpPOEntry.grid(row = 5, column = 1, columnspan = 2, sticky = "nsew")
    Statuslabel = tk.Label(window, text = "Status of chamber: " + df.at[x + (6 * (bay_num - 1)),'Status Of Chamber'])
    Statuslabel.grid(row = 6, column = 1, columnspan = 2, sticky = "nsew")
    Passdownlabel = tk.Label(window, text = "Passdown issues: " + str(df.at[x + (6 * (bay_num - 1)),'Passdown Issues']))
    Passdownlabel.grid(row = 7, column = 1, columnspan = 2, sticky = "nsew")
    StartDatelabel = tk.Label(window, text = "Start Date: " + str(df.at[x + (6 * (bay_num - 1)),'START Date']))
    StartDatelabel.grid(row = 8, column = 1, sticky = "nsew")
    PortDayslabel = tk.Label(window, text = "Port days: " + str(df.at[x + (6 * (bay_num - 1)),'Port Days']))
    PortDayslabel.grid(row = 8, column = 2, sticky = "nsew")
    #Change check if none to check if xxxxx
    #if chamberPO is not None:
    chUpdateQNButton = tk.Button(window, text = "Update chamber QN's", command = (lambda: updateQN(chamberPO, chamber, True, 0)))
    chUpdateQNButton.grid(row = 9, column = 1, sticky = "nsew")
    chViewQNButton = tk.Button(window, text = "View open chamber QN's", command = (lambda: viewQN(chamber, True, 0)))
    chViewQNButton.grid(row = 9, column = 2, sticky = "nsew")
    ctvUpdateQNButton = tk.Button(window, text = "Update CTV QN's", command = (lambda: updateQN(chamber.ctvPO, chamber, True, 1)))
    ctvUpdateQNButton.grid(row = 10, column = 1, sticky = "nsew")
    ctvViewQNButton = tk.Button(window, text = "View open CTV QN's", command = (lambda: viewQN(chamber, True, 1)))
    ctvViewQNButton.grid(row = 10, column = 2, sticky = "nsew")
##    gpPO = chambers[x + (6 * (bay_num - 1))].gpPO
##    if not pd.isna(gpPO):
##        gpQNButton = tk.Button(window, text = "View gas panel QN's", command = (lambda: openqn(gpPO)))
##        gpQNButton.pack()
    if chamberPO is not None:
        tlc_button = tk.Button(window, text = "Add TLC to chamber", command = (lambda: opentlc(chamberPO)))#chambers[x + (6 * (bay_num - 1))].po)))
        tlc_button.grid(row = 11, column = 1, sticky = "nsew")
    checkUserTLCButton = tk.Button(window, text = "View today's TLC", command = (lambda: openUserTLC()))
    checkUserTLCButton.grid(row = 11, column = 2, sticky = "nsew")

    checkRackStatusButton = tk.Button(window, text = "Check ERack for this system", command = (lambda: printStatus(system)))
    checkRackStatusButton.grid(row = 12, column = 1, columnspan = 2, sticky = "nsew")#pack()
    
    updateSystemNumEntry = tk.Entry(window)
    updateSystemNumEntry.grid(row = 13, column = 1, sticky = "nsew")
    updateSystemNumButton = tk.Button(window, text = "Update System #", command = (lambda: setNewPO(chamber, updateSystemNumEntry, systemEntry, 0)))
    updateSystemNumButton.grid(row = 13, column = 2, sticky = "nsew")
    
    
    updateChPOEntry = tk.Entry(window)
    updateChPOEntry.grid(row = 14, column = 1, sticky = "nsew")
    updateChPOButton = tk.Button(window, text = "Update Chamber PO", command = (lambda: setNewPO(chamber, updateChPOEntry, POEntry, 1)))
    updateChPOButton.grid(row = 14, column = 2, columnspan = 2, sticky = "nsew")
    updateCTVPOEntry = tk.Entry(window)
    updateCTVPOEntry.grid(row = 15, column = 1, sticky = "nsew")
    updateCTVPOButton = tk.Button(window, text = "Update CTV PO", command = (lambda: setNewPO(chamber, updateCTVPOEntry, ctvPOEntry, 2)))
    updateCTVPOButton.grid(row = 15, column = 2, columnspan = 2, sticky = "nsew")

    updateGPPOEntry = tk.Entry(window)
    updateGPPOEntry.grid(row = 16, column = 1, sticky = "nsew")
    updateGPPOButton = tk.Button(window, text = "Update Gas Panel PO", command = (lambda: setNewPO(chamber, updateGPPOEntry, gpPOEntry, 3)))
    updateGPPOButton.grid(row = 16, column = 2, columnspan = 2, sticky = "nsew")
    createY7Button = tk.Button(window, text = "Create Y7", command = (lambda: createQN(chamberPO, chamber.system, "Y7")))
    createY7Button.grid(row = 17, column = 1, sticky = "nsew")
    createY8Button = tk.Button(window, text = "Create Y8", command = (lambda: createQN(chamberPO, chamber.system, "Y8")))
    createY8Button.grid(row = 17, column = 2, sticky = "nsew")
    iomsButton = tk.Button(window, text = "Go to iOMS", command = (lambda: webbrowser.get(chrome).open_new_tab("http://ioms/MFG/ModuleStatus?PO=" + chamberPO + "#!/")))
    iomsButton.grid(row = 18, column = 1, columnspan = 4, sticky = "nsew")
    create3DButton = tk.Button(window, text = "Create new 3D form", command = (lambda: webbrowser.get(chrome).open_new_tab("http://sppartner/sites/Global3D/Lists/VMORevision/Item/newifs.aspx")))
    create3DButton.grid(row = 19, column = 1, columnspan=4, sticky = "nsew")
    
    quit_buttonGeneric = tk.Button(window, text = "quit", command = window.destroy)
    quit_buttonGeneric.grid(row = 20, column = 0, columnspan = 4, sticky = "nsew")#pack(side = "left")
    window.focus_set()                                                        
    window.grab_set()
#http://dca-wb-263/QM/QM/CreateQN?prodId=1552105&qntype=Y8&slotno=B01487&plant=4070&source=PROMPT

def change_bay(root, chamber_image, new_bay, entry, active_buttons, currentBayLabel):
    global chamber_locations, bay_num_str, bay_num
    entry.delete('0', 'end')
    try:
        bay_num = int(new_bay)
        bay_num_str = new_bay
    except ValueError:
        return
    delete_buttons()
    chamber_locations = findChamberLocations(bay_num)
    create_buttons(root, chamber_image, chamber_locations, active_buttons)
    currentBayLabel['text'] = "Now viewing Bay " + bay_num_str


def delete_buttons():
    global active_buttons, chamber_locations
    x = 0
    for x in active_buttons:
        x.destroy()
    active_buttons.clear()

########################NEED TO RESTORE FOCUS TO SECOND WINDOW AFTER CLOSING THIRD
def updateAllBayQNs(bay_num):
    options = Options()
    options.headless = True # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    for x in range(6):
        chamber = chambers[x + (6 * (bay_num - 1))]
        chamberPO = chamber.chPO
        updateBayQN(driver, chamberPO, chamber)
    driver.quit()
#need to add ctv collecting
def updateBayQN(driver, po, ch):
    driver.get("http://dca-wb-263/QM/QM/ViewQN?SlotNum=&ProdOrder=" + po)
    try:
        e = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "ui-grid-row")))
        ch.QNs.clear()
        rows = driver.find_elements_by_class_name("ui-grid-row")
        for r in rows:
            qnnum = r.find_element_by_css_selector('a.ng-binding').text
            qnType = r.find_element_by_class_name('ui-grid-coluiGrid-000K').text
            shortText = r.find_element_by_class_name('ui-grid-coluiGrid-000L').text
            status = not r.find_element_by_xpath('.//input[@type="checkbox"]').is_selected()
            partNum = r.find_element_by_class_name('ui-grid-coluiGrid-000M').text
            if status and (qnType == "Y8"):
                print(partNum)
                PN = fnmatch.filter(partNum.split(), '????-?????')
                print(PN)
                if PN:
                    partNum = PN
                    ch.QNs.append(QN(qnnum, qnType, shortText, status, partNum))
                    continue
                else:
                    PNReg = re.compile(r'\d{4}-\d{5}')
                    foundPN = PNReg.search(shortText)
                    if foundPN:
                        print(qnnum)
                        print(foundPN.group())
                        partNum = foundPN.group()
                        ch.QNs.append(QN(qnnum, qnType, shortText, status, partNum))
                        continue
            ch.QNs.append(QN(qnnum, qnType, shortText, status))
    except:
        ch.QNs.clear()
############################################################################################################3
def updateQN(po, ch, viewQNs, qnIdx):
    options = Options()
    options.headless = True # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    driver.get("http://dca-wb-263/QM/QM/ViewQN?SlotNum=&ProdOrder=" + po)
    try:
        e = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "ui-grid-row")))
        ch.QNs.clear()
        rows = driver.find_elements_by_class_name("ui-grid-row")
        for r in rows:
            posPartNums = []
            qnnum = r.find_element_by_css_selector('a.ng-binding').text
            qnType = r.find_element_by_class_name('ui-grid-coluiGrid-000K').text
            shortText = r.find_element_by_class_name('ui-grid-coluiGrid-000L').text
            status = not r.find_element_by_xpath('.//input[@type="checkbox"]').is_selected()
            partNum = r.find_element_by_class_name('ui-grid-coluiGrid-000M').text
            if qnIdx == 0:
                if status:
                    print(partNum)
                    PN = fnmatch.filter(partNum.split(), '????-?????')
                    print(PN)
                    if PN:
                        partNum = PN
                        ch.QNs.append(QN(qnnum, qnType, shortText, status, partNum))
                        continue
                    else:
                        PNReg = re.compile(r'\d{4}-\d{5}')
                        foundPN = PNReg.search(shortText)
                        if foundPN:
                            print(qnnum)
                            print(foundPN.group())
                            partNum = foundPN.group()
                            ch.QNs.append(QN(qnnum, qnType, shortText, status, partNum))
                            continue
                ch.QNs.append(QN(qnnum, qnType, shortText, status))
            elif qnIdx == 1:
                ch.ctvQNs.append(QN(qnnum, qnType, shortText, status))
            else:
                ch.gpQNs.append(QN(qnnum, qnType, shortText, status))
        driver.quit()
    except:
        driver.quit()
        ch.allQNs[qnIdx].clear()
    if viewQNs:
        viewQN(ch, True, qnIdx)

def createQN(po, system, qnType):
    x = system.find("-")
    system = system[0:x]
    print(system)
    webbrowser.get(chrome).open_new_tab("http://dca-wb-263/QM/QM/CreateQN?prodId=" + po + "&qntype=" + qnType +"&slotno=" + system + "&plant=4070&source=PROMPT")

def openQN(qn):
    webbrowser.get(chrome).open_new_tab("https://epvpwd.amat.com:8065/com.amat.irj.portal?app=ChgQaNotif?RIWO00-QMNUM=0000" + qn)

def viewQN(ch, onlyOpen, qnIdx):
    status = ""
    if onlyOpen:
        status = "Open QN's for " + ch.system + ":\n------------------------------------------------------------\n"
    else:
        status = "All QN's for " + ch.system + ":\n-----------------------------------------------------------------\n"
    #status = ch.QNStatus(onlyOpen, status, qnIdx)
    window = tk.Toplevel(root)
    qns_label = tk.Label(window, text = status)
    qns_label.pack()
    openQNs = ch.openQNs(qnIdx)
    #numOpen = ch.numOpenQNs(qnIdx)
    numOpen = len(openQNs)
    if numOpen != 0:
        qnLinks = []
        texts = []
        #print(numOpen)
        for i in range(numOpen):
            #status2 = ch.QNStatus(onlyOpen, status, qnIdx)
            texts.append(tk.Text(window, height = 2, borderwidth = 0))
            qnLinks.append(texts[i])
            qnLinks[i].tag_configure("center", justify='center')
            qnLinks[i].tag_add("center", "1.0", "end")
            qnLinks[i].configure(bg=window.cget('bg'), relief="flat")
            qnLinks[i].pack()
            hyperlink = HyperlinkManager(qnLinks[i])
            qnLinks[i].insert(INSERT, openQNs[i].Type + " ")
            qnLinks[i].insert(INSERT, openQNs[i].QNNum, hyperlink.add(lambda x = openQNs[i].QNNum: openQN(x)))
            qnLinks[i].insert(INSERT, " - " + openQNs[i].Desc.upper())
            qnLinks[i].configure(state="disabled")
            if openQNs[i].partNum and (openQNs[i].Type == "Y8"):
                ltsText = tk.Text(window, height = 1, borderwidth =0)
                ltsText.tag_configure("center", justify='center')
                ltsText.tag_add("center", "1.0", "end")
                ltsText.configure(bg=window.cget('bg'), relief="flat")
                ltsText.pack()
                ltsButton = tk.Button(window, text="Check last scanned", command = (lambda text = ltsText: getLastScanned(ch.chPO, openQNs[i], text)))
                ltsButton.pack()
                ltsText = tk.Text(window, height = 1, borderwidth =0)

    else:
        tk.Label(window, text = "No QN data").pack()
    #print(i)
    quit_button_qns = tk.Button(window, text = "quit", command = window.destroy)
    quit_button_qns.pack(side = "left")
    window.focus_set()                                                        
    window.grab_set()



def getLastScanned(po, qn, text):
    options = Options()
    options.headless = True # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    driver.get("http://dca-app-833/LTSWeb/LTSPACKAGE")
    #try:
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="PO"]')))
    poEntry = driver.find_element_by_xpath('//*[@id="PO"]')
    poEntry.send_keys(po)
    partNumEntry = driver.find_element_by_xpath('//*[@id="PN"]')
    partNumEntry.send_keys(qn.partNum)
    driver.find_element_by_xpath('//*[@id="btnSearch"]/i').click()
    time.sleep(5)
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "ui-grid-row")))
        row = driver.find_element_by_class_name("ui-grid-row")
        #matNum = row.find_element_by_class_name('ui-grid-coluiGrid-000A').text
        lastSeen = row.find_element_by_class_name("ui-grid-coluiGrid-000C").text
        loc = row.find_element_by_class_name("ui-grid-coluiGrid-000F").text
        driver.quit()
        print("Last seen: " + lastSeen + " in " + loc)
        text.delete('1.0', END)
        text.insert(INSERT, ("Last seen: " + lastSeen + " in " + loc))
    except:
        driver.quit()
        text.delete('1.0', END)
        text.insert(INSERT, "No Data Found")


def opentlc(po):
    options = Options()
    options.headless = False # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    driver.get("http://ioms/MFG/ModuleStatus?PO=" + po + "#!/laborcosting")
    time.sleep(15)
    but = WebDriverWait(driver, 35).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#laborcosting > div > div > div > div > div:nth-child(6) > div.col-md-5.col-sm-6 > div > button:nth-child(1) > span')))
    print("is clickable")
    time.sleep(2)
    but.click()
    op = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#addTestRecord > div > div > form > div.modal-body.mdbody > div:nth-child(1) > div.col-md-4 > select')))
    time.sleep(2)
    op.click()
    modtest = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#addTestRecord > div > div > form > div.modal-body.mdbody > div:nth-child(1) > div.col-md-4 > select > option.ng-binding.ng-scope')))
    time.sleep(2)
    modtest.click()
    atype = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#addTestRecord > div > div > form > div.modal-body.mdbody > div:nth-child(3) > div.col-md-4 > select')))
    time.sleep(2)
    atype.click()
    hrs = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#addTestRecord > div > div > form > div.modal-body.mdbody > div:nth-child(7) > div.col-md-4 > div > input:nth-child(1)')))
    time.sleep(2)
    hrs.send_keys("8")
    time.sleep(5)
    driver.quit()


def openUserTLC():
     webbrowser.get(chrome).open_new_tab("http://dca-wb-263/PROMPT/ToolLaborCost/BiWeeklyTLC")


#def updateLastTestRan(bay_num_str, port):
    # TODO: implement 

def updateSystemNum(num, window):
    print("changed to " + num)
    window.destroy()
    
def setNewPO(ch, new, old, poType):
    old.configure(state="normal")
    old.delete('1.0', END)
    po = new.get()
    new.delete('0', 'end')
    t = ("Chamber PO: " if poType == 1 else ("CTV PO: " if poType == 2 else "GP PO: ")) + po
    old.insert(1.0, t)
    old.tag_configure("center", justify='center')
    old.tag_add("center", "1.0", "end")
    old.configure(state="disabled")
    manualPOsFile = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx'
    manualPOs = load_workbook(manualPOsFile)
    ws = manualPOs.active
    found = False
    rows = list(ws.rows)
    r = rows[ch.portIdx + 2]
    if r[1].value == ch.system:
        found = True
        r[poType + 2].value = po
        manualPOs.save(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
    if found == False:
        ws.cell(column=2, row=ch.portIdx + 2, value=ch.system)
        ws.cell(column=(poType + 2), row=ch.portIdx + 2, value=po)
        manualPOs.save(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
    if poType == 0:
        ch.system = po
    elif poType == 1:
        ch.chPO = po
    elif poType == 2:
        ch.ctvPO = po
    else:
        ch.gpPO = po

def openUpdateSystemNumWindow():
    window = tk.Toplevel(root)
    window.geometry('+%d+%d' % (690, 100))
    updateSystemNumEntry = tk.Entry(window)
    updateSystemNumEntry.pack()
    updateSystemNumButton = tk.Button(window, text = "Update System #", command = (lambda: updateSystemNum(updateSystemNumEntry.get(), window)))
    updateSystemNumButton.pack()
    window.focus_set()
    window.grab_set()

def refreshRackStatus():
    fname = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\SUPERVISOR PASSDOWN\EQRK Status.xls'
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx', FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    wb.Close()                               #FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    del excel

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)
    
############################################################ MAIN PROGRAM BEGINS ########################################################################
print(datetime.today())
skip = True

rackStatus = load_workbook(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx')
rackStatusSheet = rackStatus["Sheet1"]
manualPOsFile = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx'
passdownPath = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\SUPERVISOR PASSDOWN\LEADS Passdown\LEADS PASSDOWN*.xlsx'
crossoverPath = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\SUPERVISOR PASSDOWN\( DTF Checklists for Systems )\(TEST QUEUE )\TEST QUEUE*.xlsx'
manualPOPath = (r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
chrome = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"
crossoverFile = getLatestFile(crossoverPath)
print(crossoverFile)
crossoverPath = crossoverPath.replace("TEST QUEUE*.xlsx", crossoverFile)
passdownFile = getLatestFile(passdownPath)
print(passdownFile)
passdownPath = passdownPath.replace("LEADS PASSDOWN*.xlsx", passdownFile)
chamber_image_file = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\chamber.png'
bay_image_file = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\baydrawingNEW.png'
chambers = []
cells = []
active_buttons = []
chamber_locations = [1,1,1,1,1,1]
ports = ["A","B","C","D","E","F"]

getPriorityColors(passdownPath, cells)

data = pd.read_excel(crossoverPath, sheet_name= 'QUEUE', usecols = 'F:H', dtype=str, skiprows = 4)
dfCrossover = pd.DataFrame(data)
data = pd.read_excel(manualPOPath, sheet_name = 'Sheet1', usecols = 'A:E', dtype = str)
manualDF = pd.DataFrame(data)

bay_num_str = "8"
bay_num = 8

data = pd.read_excel(passdownPath)
df = pd.DataFrame(data, columns= ['Bay ','System #', 'Status Of Chamber', 'Passdown Issues','START Date','Port Days'])

getPOs(chambers)
chamber_locations = findChamberLocations(bay_num)
print(chamber_locations)
#manualRows = 
root = tk.Tk()
root.title("Bay Status")
w = 1154
h = 881
ws = root.winfo_screenwidth() # width of the screen
hs = root.winfo_screenheight() # height of the screen
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
root.geometry('%dx%d+%d+%d' % (w, h, x, y-40))
canvas = Canvas(root, width=1154, height=881)
bay_image = ImageTk.PhotoImage(file = bay_image_file)
currentBayLabel = tk.Label(root, text = "Now viewing Bay " + bay_num_str)
currentBayLabel.grid(row = 0, column = 0, sticky = "w")
new_bay_entry = tk.Entry(root)
new_bay_entry.grid(row = 0, column = 1, sticky = "e")#pack()
refresh_button = tk.Button(root, text = "Change bay", command = (lambda: change_bay(root, chamber_image, new_bay_entry.get(), new_bay_entry, active_buttons, currentBayLabel)))#locations
refresh_button.grid(row = 0, column = 2, sticky = "nsew")#pack()
rackStatusEntry = tk.Entry(root)
rackStatusEntry.grid(row = 1, column = 1, sticky = "e")#pack()
rackStatusButton = tk.Button(root, text = "Check ERack status for system", command = (lambda: printStatus(rackStatusEntry.get().upper())))
rackStatusButton.grid(row = 1, column = 2, sticky = "nsew")#pack()
rackRefreshButton = tk.Button(root, text = "Refresh ERack status", command = (lambda: refreshRackStatus()))
rackRefreshButton.grid(row = 1, column = 3, sticky = "w")#pack()
updateAllBayQNsButton = tk.Button(root, text = "Update QN's for whole bay", command = (lambda: updateAllBayQNs(bay_num)))
updateAllBayQNsButton.grid(row = 2, column = 2)#pack()
canvas.create_image(575,450, image = bay_image)
canvas.grid(row = 3, column = 0, columnspan = 4)#pack()
chamber_image = PhotoImage(file = chamber_image_file)

create_buttons(root, chamber_image, chamber_locations, active_buttons)
    

root.mainloop()

