from openpyxl import Workbook
from openpyxl import load_workbook
import win32com.client as win32

# Only need to run this once to create a copy of eqrk status saved in a newer format.
# After the copy is made only need to run this to update the copy

##fname = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\SUPERVISOR PASSDOWN\EQRK Status.xls'
##excel = win32.gencache.EnsureDispatch('Excel.Application')
##wb = excel.Workbooks.Open(fname)
##
##wb.SaveAs(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx', FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
##wb.Close()                               #FileFormat = 56 is for .xls extension
##excel.Application.Quit()
##del excel

#Looks for all ready racks for given system, needs updated eqrk status to work properly


wb = load_workbook(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx')
ws = wb["Sheet1"]
def printStatus(system):
    rows = ws.iter_rows(ws.max_row - 100, ws.max_row)
    for row in rows:
        if row[0].value == system:
            for x in range(8,12):
                if "NA" in str(row[x].value):
                    print("There is no EQRK #" + str(x - 7) + "\n")
                elif row[x].value == None:
                    print("EQRK #" + str(x - 7) + " is not ready, needed for:")
                    x = str(row[x].comment)
                    y = x.find('CH')
                    z = x.find('by')
                    print(x[y:z-1] + "\n")
                else:
                    print("EQRK #" + str(x - 7) + " is complete for:")
                    #May want to get completed date: row[x].value
                    x = str(row[x].comment)
                    y = x.find('CH')
                    z = x.find('by')
                    print(x[y:z-1] + "\n")
            try:
                print("Comments: " + row[12].value)
            except:
                pass
while(input != "STOP"):
    system = input("Enter system: ").upper()
    printStatus(system)
