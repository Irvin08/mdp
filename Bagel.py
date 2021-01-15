#Bagel Beta release
#Irvin Carrillo 170821
#12/21/20
#Tool to gather and track relevant data
#in MDP Final Test
#Add last scanned to copy qns to clipboard button


import pandas as pd
from datetime import datetime, timedelta
from datetime import timezone
from PIL import Image, ImageDraw
import tkinter as tk
from tkinter import *
from tkinter import ttk
from PIL import ImageTk
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import webbrowser
import win32com.client as win32
import os.path
from os import path
import glob
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
import fnmatch
import time
import random
import json
import math
import re
import configparser


##################################################################################
#                                                                                #
#                                    Classes                                     #
#                                                                                #
##################################################################################
class Bay:
    def __init__ (self, bayNumber, portNumber):
        self.bayNumber = bayNumber
        self.portNumber = portNumber
        self.chambers = []



class Chamber:
    def __init__ (self, port, system, chamberPO, ctvPO, gpPO, chType, portIdx):
        self.bay = int(re.sub("\D", "", port))
        self.port = port
        self.system = system
        self.chPO = chamberPO
        self.ctvPO = ctvPO
        self.gpPO = gpPO
        self.chType = chType
        self.portIdx = portIdx
        self.gpQNs = []
        self.ctvQNs = []
        self.QNs = []
        self.InspLots = []
        self.ESWs = []
        self.ctvInspLots = []
        self.gpInspLots = []
        self.allQNs = [self.QNs, self.ctvQNs, self.gpQNs]
        self.allInspLots = [self.InspLots, self.ctvInspLots, self.gpInspLots]
        self.currentMayanSection = ""
        self.currentMayanTest = ""
        self.mayanUpdateTimestamp = ""
        self.buildPercent = "-"
        self.color = None

    def __str__(self):
        return 'System #:{}, PO #: {}'.format(self.system,self.chPO)

    def print(self):
        print(self.system + " " + self.chPO + " " + self.chType)

    def numOpenQNs(self, qnIdx):
        openQNs = 0
        for q in self.allQNs[qnIdx]:
            if q.isOpen:
                openQNs = openQNs + 1
        return openQNs

    def openQNs(self, qnIdx):
        qnList = []
        for q in self.allQNs[qnIdx]:
            if q.isOpen:
                qnList.append(q)
        return qnList

    def openInspLots(self, Idx):
        ilList = []
        for i in self.allInspLots[Idx]:
            if i.isOpen:
                ilList.append(i)
        return ilList
    
    def QNStatus(self, onlyOpen, status, qnIdx):
        if not self.allQNs[qnIdx]:
            return "There are no QN's\n"
        allClosed = True
        if onlyOpen:
            for q in self.allQNs[qnIdx]:
                if q.isOpen:
                    status = status + q.print()
                    allClosed = False
        else:
            for q in self.allQNs[qnIdx]:
                if q.isOpen:
                    allClosed = False
                status = status + q.print()
        if onlyOpen and allClosed:
            status = "There are no open QN's\n"
        return status

#Currently unused, here in case needed when merging extended bays
class Port:
    def __init__ (self, portName, occupied, chamber, chamberPO, chamberType):
        self.portName = portName
        self.occupied = occupied
        self.chamber = chamber
        self.chamberPO = chamberPO
        self.chamberType = chamberType

    def info(self):
        return self.portName + '|' + self.chamber + '|' +self.chamberPO + '|' + self.chamberType + '\n'
    
    def print(self):
        print(self.portName + ' ' + self.chamber + ' ' +self.chamberPO)


class QN:
    def __init__ (self, QNNum, Type, Desc, isOpen, partNum = None):
        self.QNNum = QNNum
        self.Type = Type
        self.Desc = Desc
        self.isOpen = isOpen
        self.partNum = partNum
        self.dateClosed = None
        self.creator = None
        self.lastScanned = None
        self.lastUpdateTimestamp = None
        
    def isOpen(self):
        return self.isOpen

    def print(self):
        return (self.Type + " " + self.QNNum + " - " + self.Desc.upper() + ("\n" if self.isOpen else " | CLOSED\n"))


class InspLot:
    def __init__ (self, lotNum, Desc, Status, isOpen):
        self.lotNum = lotNum
        self.Desc = Desc
        self.Status = Status
        self.isOpen = isOpen

    def isOpen(self):
        return self.isOpen()

    def print(self):
        return ("InspLot " + self.lotNum + " - " + self.Desc + ("\n" if self.isOpen else (" | " + self.Status + "\n")))

#Used to create QN/InspLot hyperlinks 
class HyperlinkManager:
    def __init__(self, text):
        self.text = text
        self.text.tag_config("hyper", foreground="blue", underline=1)
        self.text.tag_bind("hyper", "<Enter>", self._enter)
        self.text.tag_bind("hyper", "<Leave>", self._leave)
        self.text.tag_bind("hyper", "<Button-1>", self._click)
        self.reset()

    def reset(self):
        self.links = {}

    def add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hyper-%d" % len(self.links)
        self.links[tag] = action
        return "hyper", tag

    def _enter(self, event):
        self.text.config(cursor="hand2")

    def _leave(self, event):
        self.text.config(cursor="")

    def _click(self, event):
        for tag in self.text.tag_names(CURRENT):
            if tag[:6] == "hyper-":
                self.links[tag]()
                return


##################################################################################
#                                                                                #
#                                 System Functions                               #
#                                                                                #
##################################################################################
#Returns name of most recent file in path
def getLatestFile(path):
    list_of_files = glob.glob(path)
    if not list_of_files:
        return None
    latest_file = max(list_of_files, key=os.path.getctime)
    _, filename = os.path.split(latest_file)
    return filename


#Copy Erack Status sheet to working folder and convert to .xlsx format
def refreshRackStatus():
    fname = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\SUPERVISOR PASSDOWN\EQRK Status.xls'
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)
    #FileFormat = 51 is for .xlsx extension, FileFormat = 56 is for .xls extension
    wb.SaveAs(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx', FileFormat = 51)
    wb.Close()                               
    excel.Application.Quit()
    del excel


#Used for driver creation
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


#Copy Item(s) to clipboard
def copyQNsToClipboard(ch):
    result = ""
    openQNs = ch.openQNs(0)
    openILs = ch.openInspLots(0)
    for qn in openQNs:
        result = result + qn.print()
    for il in openILs:
        result = result + il.print()
    r = Tk()
    r.withdraw()
    r.clipboard_clear()
    r.clipboard_append(result)
    r.update()
    r.destroy()


#Create an instance of chrome driver
def createDriver(hide):
    options = Options()
    options.headless = hide # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    return driver


##################################################################################
#                                                                                #
#                                Read/Write Data                                 #
#                                                                                #
##################################################################################
#Find PO's for a lst of chamber by looking through "Test Queue" file maintained by MDP Test DTF
def getPOs(chambers):
    global dfCrossover
    global df
    global dfManual
    manualPOs = load_workbook(manualPOsFile)
    ws = manualPOs.active
    rows = list(ws.rows)
    global manualRows
    for x in range(0, 66):
        port = df.at[x, "Bay "]
        port = port.replace("Bay ", "").strip()
        skip = False
        done = False
        y = df.at[x, 'System #']
        y = y.replace(" ", "")        
        if "EMPTY" in y:
            chambers.append(Chamber(port, "XXXXX-X", "XXXXX", "XXXXX", "XXXXX", "XXXXX", x))
            skip = True
        if not skip:
            z = y.find('-')
            y = y[0 : z + 2]
            #print(y)
            if "D01" in y:
                try:
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '].astype(str).str.contains(y)]
                    chambers.append(Chamber(port, y, r.values[0,1], "XXXXX", "XXXXX", r.values[0,2], x))
                except IndexError:
                    #y = y.replace("D01", "")
                    temp = y.replace("D01", "")
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '].astype(str).str.contains(temp)]
                    chambers.append(Chamber(port, y, r.values[0,1], "XXXXX", "XXXXX", r.values[0,2], x))
            else:
                try:
                    r = dfCrossover.loc[dfCrossover['Slot /Sys - Ch# '].astype(str).str.contains(y)]
                    po = r.values[0,1]
                    chtype = r.values[0,2]
                    found = True
                except IndexError:
                    pass
            r = rows[x + 1]               
            if r[1].value == y:
                init = [port, y, po, None, None, chtype, x]
                z = [cell.value for cell in r]
                for i in range(1,6):
                    if z[i]:
                        init[i] = z[i] 
                chambers.append(Chamber(*init))
                done = True
            if not done:
                chambers.append(Chamber(port, y, po, "XXXXX", "XXXXX", chtype, x)) if found else chambers.append(Chamber(port, y, "XXXXX", "XXXXX", "XXXXX", "XXXXXX", x))


#Change or add relevant PO, this does not write to "Test Queue" file 
def setNewPO(ch, new, old, poType):
    old.configure(state="normal")
    old.delete('1.0', END)
    po = new.get()
    new.delete('0', 'end')
    t = ("" if poType == 0 else ("Chamber PO: " if poType == 1 else ("CTV PO: " if poType == 2 else "GP PO: "))) + po
    old.insert(1.0, t)
    old.tag_configure("center", justify='center')
    old.tag_add("center", "1.0", "end")
    old.configure(state="disabled")
    manualPOsFile = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx'
    manualPOs = load_workbook(manualPOsFile)
    ws = manualPOs.active
    found = False
    rows = list(ws.rows)
    r = rows[ch.portIdx + 2]
    if r[1].value == ch.system:
        found = True
        r[poType + 2].value = po
        manualPOs.save(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
    if found == False:
        ws.cell(column=2, row=ch.portIdx + 2, value=ch.system)
        ws.cell(column=(poType + 2), row=ch.portIdx + 2, value=po)
        manualPOs.save(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
    if poType == 0:
        ch.system = po
    elif poType == 1:
        ch.chPO = po
    elif poType == 2:
        ch.ctvPO = po
    else:
        ch.gpPO = po


#Get leads passdoesn color codes, not super relevant may remove
def getPriorityColors(file, chambers):
    wb = load_workbook(file, data_only = True)
    sh = wb['Lead Passdown']
    for x in range (66):
        color_in_hex = sh["A" + str(x+2)].fill.start_color.index
        if color_in_hex != 0:
            try:
                rgb = tuple(int(color_in_hex[i:i+2], 16) for i in ( 2, 4, 6))
                chambers[x].color = rgb
                #cells.append(rgb)
            except TypeError:
                rgb = (255, 255, 255)
                chambers[x].color = rgb
                #cells.append(rgb)
        else:
            chambers[x].color = (255, 255, 255)
            #cells.append((255, 255, 255))
    #del cells[0]


#Gets and displays erack data for given system in a new window
def printStatus(system):
    #print(system)
    rows = rackStatusSheet.iter_rows(rackStatusSheet.max_row - 200, rackStatusSheet.max_row)
    status = ""
    found = False
    for row in rows:
        if row[0].value == system:
            found = True
            for x in range(8,12):
                if "NA" in str(row[x].value):
                    #print("There is no EQRK #" + str(x - 7) + "\n")
                    status = status + "There is no EQRK #" + str(x - 7) + "\n --------------------------------------\n"
                elif row[x].value == None:
                    #print("EQRK #" + str(x - 7) + " is not ready, needed for:")
                    status = status + "EQRK #" + str(x - 7) + " is not ready, needed for:"
                    x = str(row[x].comment)
                    y = x.find(':')
                    y = x.find(':', y + 1)
                    y = y + 1
                    z = x.find('by')
                    #print(x[y:z-1] + "\n")
                    status = status + x[y:z-1] + "\n --------------------------------------\n"
                else:
                    #print("EQRK #" + str(x - 7) + " is complete for:")
                    status = status + "EQRK #" + str(x - 7) + " is complete for:"
                    #May want to get completed date: row[x].value
                    x = str(row[x].comment)
                    y = x.find(':')
                    y = x.find(':', y + 1)
                    y = y + 1
                    z = x.find('by')
                    #print(x[y:z-1] + "\n")
                    status = status + x[y:z-1] + "\n --------------------------------------\n"
            try:
                #print("Comments: " + row[12].value)
                status = status + "Comments:\n" + row[12].value + "\n"
                ##ADD CHECK IF NEXT ROW IS EMPTY, IF SO NEED TO GET NEXT COMMENT 
            except:
                pass
    if not found:
        status = "EQRK not started"
        #print(status)
    window = tk.Toplevel(root)
    window.geometry('+%d+%d' % (690, 100))
    rackEntry = tk.Text(window, font = "Helvetica 12", width = 60, borderwidth = 0)
    rackEntry.insert(END, status)
    rackEntry.tag_configure("center", justify='center')
    rackEntry.tag_add("center", "1.0", "end")
    rackEntry.configure(state="disabled")
    rackEntry.configure(bg=root.cget('bg'), relief="flat")
    rackEntry.pack()
    quit_button_rack = tk.Button(window, text = "quit", command = window.destroy)
    quit_button_rack.pack(side = "left")
    window.focus_set()                                                        
    window.grab_set()


##################################################################################
#                                                                                #
#                                  GUI Drawing                                   #
#                                                                                #
##################################################################################
#Find what ports have a chamber
def findChamberLocations(bay):
    global chamber_locations
    if bays[bay].portNumber == 6:
        for i in range(6):
            chamber_locations[i] = 1 if bays[bay].chambers[i].system != "XXXXX-X" else 0
        chamber_locations[6:] = [0] * 2
    else:
        for i in range(8):
            chamber_locations[i] = 1 if bays[bay].chambers[i].system != "XXXXX-X" else 0
    return chamber_locations


#Create the chamber buttons basd on what ports have a chamber
def create_buttons(root, chamber_image, chamber_locations, active_buttons):
    global bay_num
    if bays[bay_num].portNumber == 6:
        if chamber_locations[0] == 1:
            portA_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 0)), anchor = "w")
            portA_button_window = canvas.create_window(890,20, anchor= "nw", window = portA_button)
            active_buttons.append(portA_button)

        if chamber_locations[1] == 1:
            portB_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 1)), anchor = "w")
            portB_button_window = canvas.create_window(890,255, anchor= "nw", window = portB_button)
            active_buttons.append(portB_button)

        if chamber_locations[2] == 1:
            portC_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 2)), anchor = "w")
            portC_button_window = canvas.create_window(890,490, anchor= "nw", window = portC_button)
            active_buttons.append(portC_button)

        if chamber_locations[3] == 1:
            portF_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 3)), anchor = "w")
            portF_button_window = canvas.create_window(130,490, anchor= "nw", window = portF_button)
            active_buttons.append(portF_button)

        if chamber_locations[4] == 1:
            portG_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 4)), anchor = "w")
            portG_button_window = canvas.create_window(130,250, anchor= "nw", window = portG_button)
            active_buttons.append(portG_button)

        if chamber_locations[5] == 1:
            portH_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 5)), anchor = "w")
            portH_button_window = canvas.create_window(130,20, anchor= "nw", window = portH_button)
            active_buttons.append(portH_button)
    else:
        if chamber_locations[0] == 1:
            portA_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 0)), anchor = "w")
            portA_button_window = canvas.create_window(890,20, anchor= "nw", window = portA_button)
            active_buttons.append(portA_button)

        if chamber_locations[1] == 1:
            portB_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 1)), anchor = "w")
            portB_button_window = canvas.create_window(890,255, anchor= "nw", window = portB_button)
            active_buttons.append(portB_button)

        if chamber_locations[2] == 1:
            portC_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 2)), anchor = "w")
            portC_button_window = canvas.create_window(890,490, anchor= "nw", window = portC_button)
            active_buttons.append(portC_button)

        if chamber_locations[3] == 1:
            portD_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 3)), anchor = "w")
            portD_button_window = canvas.create_window(890,730, anchor= "nw", window = portD_button)
            active_buttons.append(portD_button)

        if chamber_locations[4] == 1:
            portE_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 4)), anchor = "w")
            portE_button_window = canvas.create_window(130,730, anchor= "nw", window = portE_button)
            active_buttons.append(portE_button)

        if chamber_locations[5] == 1:
            portF_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 5)), anchor = "w")
            portF_button_window = canvas.create_window(130,490, anchor= "nw", window = portF_button)
            active_buttons.append(portF_button)

        if chamber_locations[6] == 1:
            portG_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 6)), anchor = "w")
            portG_button_window = canvas.create_window(130,250, anchor= "nw", window = portG_button)
            active_buttons.append(portG_button)

        if chamber_locations[7] == 1:
            portH_button = tk.Button(root, image=chamber_image, command=(lambda: create_window_Generic(bay_num, 7)), anchor = "w")
            portH_button_window = canvas.create_window(130,20, anchor= "nw", window = portH_button)
            active_buttons.append(portH_button)


#Used when changing to a different bay, deletes buttons for curent bay
def delete_buttons():
    global active_buttons, chamber_locations
    x = 0
    for x in active_buttons:
        x.destroy()
    active_buttons.clear()


def change_bay(root, chamber_image, new_bay, entry, active_buttons, currentBayLabel, canvas, sectionTexts, statusTexts, updateTimeTexts, percentTexts):
    global chamber_locations, bay_num_str, bay_num
    entry.delete('0', 'end')
    x = 0
    try:
        x = int(new_bay)
    except ValueError:
        return
    if x < 11 and x > 0:#change to max bays
        bay_num = x
        bay_num_str = new_bay
        delete_buttons()
        chamber_locations = findChamberLocations(bay_num)
        print(chamber_locations)
        create_buttons(root, chamber_image, chamber_locations, active_buttons)
        drawMayanStatus(bay_num, sectionTexts, statusTexts, updateTimeTexts)
        drawBuildStatus(bay_num, percentTexts)
        count = 0
        while(count < 8):
            try:
                create_window_Generic(bay_num, count)
                break
            except:
                count = count + 1
        currentBayLabel['text'] = "Now Viewing Bay " + bay_num_str


def drawMayanStatus(bay, sectionTexts, statusTexts, updateTimeText):
    skip = True if bays[bay_num].portNumber == 6 else False
    j = 0
    for i in range(8):
        if i in [3,4] and skip:
            canvas.itemconfig(sectionTexts[i], text = "")
            canvas.itemconfig(statusTexts[i], text = "")
            continue
        else:
            chamber = bays[bay].chambers[j]
            canvas.itemconfig(sectionTexts[i], text=chamber.currentMayanSection)
            canvas.itemconfig(statusTexts[i], text=chamber.currentMayanTest)
        j = j + 1
    canvas.itemconfig(updateTimeText, text="Mayan updated " + chamber.mayanUpdateTimestamp)
    


def drawBuildStatus(bay_num, percentTexts):
    skip = True if bays[bay_num].portNumber == 6 else False
    j = 0
    for i in range(8):
        if i in [3,4] and skip:
            canvas.itemconfig(percentTexts[i], text="-")
            continue
        else:
            chamber = bays[bay_num].chambers[j]
            canvas.itemconfig(percentTexts[i], text=(chamber.buildPercent + "%"))
            j = j + 1


#Creates left hand info window and buttons
def create_window_Generic(bay, x):
    passLabelHeight = 5
    chamber = bays[bay].chambers[x]
    chamberPO = chamber.chPO
    ctvPO = chamber.ctvPO
    gpPO = chamber.gpPO
    Portlabel = tk.Label(left,font = "Helvetica 16 bold", text = "Port: " + bay_num_str + ports[x])
    Portlabel.grid(row = 0, column = 0,columnspan=4, sticky = "nsew")
    if chamber.color == (0, 0, 0):
        chamber.color = (255, 255, 255)
    system = chamber.system
    systemEntry = tk.Text(left, font = "Helvetica 30 bold", height = 1, width = 16, borderwidth = 0)
    systemEntry.insert(1.0, system)
    systemEntry.configure(state="disabled")
    hawtnessRGB = ("#%02x%02x%02x" % chamber.color)
    systemEntry.configure(bg=(hawtnessRGB if hawtnessRGB != "#ffffff" else root.cget('bg')), relief="flat")
    systemEntry.grid(row = 2, column = 0,columnspan=4,sticky = "nsew")
    system = system[0 : system.find("-")]
    chTypeText = chamber.chType if chamber.chType is not None else "No chamber type found"
    chTypeEntry = tk.Text(left, font = "Helvetica 12", height = 1, width = 25, borderwidth = 0)
    chTypeEntry.insert(1.0, "CH Type: " + chTypeText)
    chTypeEntry.tag_configure("center", justify='center')
    chTypeEntry.tag_add("center", "1.0", "end")
    chTypeEntry.configure(state="disabled")
    chTypeEntry.configure(bg=root.cget('bg'), relief="flat")
    chTypeEntry.grid(row = 3, column = 0,columnspan=2, sticky = "nsew")
    POEntry = tk.Text(left, font = "Helvetica 12", height = 1, width = 25, borderwidth = 0)
    if chamberPO is None:
        POEntry.insert(1.0, "PO is not updated")
    else:
        POEntry.insert(1.0, "Chamber PO: " + chamberPO)
    POEntry.tag_configure("center", justify='center')
    POEntry.tag_add("center", "1.0", "end")
    POEntry.configure(state="disabled")
    POEntry.configure(bg=root.cget('bg'), relief="flat")
    POEntry.grid(row = 3, column = 2,columnspan=2, sticky = "nsew")
    ctvPOEntry = tk.Text(left,font="Helvetica 12", height = 1, width =25, borderwidth = 0)
    ctvText = "CTV PO is not updated" if (ctvPO == "XXXXX" or ctvPO == None) else "CTV PO: " + chamber.ctvPO
    ctvPOEntry.insert(1.0, ctvText)
    ctvPOEntry.tag_configure("center", justify = 'center')
    ctvPOEntry.tag_add("center", "1.0", "end")
    ctvPOEntry.configure(state="disabled")
    ctvPOEntry.configure(bg=root.cget('bg'), relief="flat")
    ctvPOEntry.grid(row = 4, column = 0, columnspan=2, sticky = "nsew")
    gpPOEntry = tk.Text(left,font="Helvetica 12", height = 1, width = 25, borderwidth = 0)
    gpText = "No Gas Panel PO available" if (gpPO == "XXXXX" or gpPO == None) else "GP PO: " + chamber.gpPO
    gpPOEntry.insert(1.0, gpText)
    gpPOEntry.tag_configure("center", justify = 'center')
    gpPOEntry.tag_add("center", "1.0", "end")
    gpPOEntry.configure(state="disabled")
    gpPOEntry.configure(bg=root.cget('bg'), relief="flat")
    gpPOEntry.grid(row = 4, column = 2, columnspan = 2, sticky = "nsew")
    firstSeparator = ttk.Separator(left, orient = HORIZONTAL)
    firstSeparator.grid(row=5,column=0,columnspan=4,rowspan=2,sticky="nsew" )
    ##############################################################################
    data = df.loc[df['System #'].astype(str).str.contains(chamber.system)]
    chStatus = data.values[0][2]
    statusEntry = tk.Text(left,font="Helvetica 10", height = 3, width = 25, borderwidth = 0, wrap=WORD)
    statusText = "Status: " + chStatus
    statusEntry.insert(1.0, statusText)
    statusEntry.tag_configure("center", justify = 'center')
    statusEntry.tag_add("center", "1.0", "end")
    statusEntry.configure(state="disabled")
    statusEntry.configure(bg=root.cget('bg'), relief="flat")
    statusEntry.grid(row = 7, column = 0, columnspan = 3, sticky = "nsew")
    PortDayslabel = tk.Label(left,font="Helvetica 10", text = "Port days: " + str(data.values[0][5]))
    PortDayslabel.grid(row = 7, column = 3, sticky = "nsew")
    chPassdown = data.values[0][3]
    Passdownlabel = tk.Label(left, font = "Helvetica 10", height = 2, text = "Passdown issues: " + chPassdown)
    Passdownlabel.grid(row = 8, column = 0, columnspan = 4, sticky = "nsew")
    passdownEntry = tk.Text(left,font="Helvetica 10", height = 4, width = 25, borderwidth = 0, wrap=WORD)
    passdownText = "Issues: " + chPassdown
    passdownEntry.insert(1.0, passdownText)
    passdownEntry.tag_configure("center", justify = 'center')
    passdownEntry.tag_add("center", "1.0", "end")
    passdownEntry.configure(state="disabled")
    passdownEntry.configure(bg=root.cget('bg'), relief="flat")
    passdownEntry.grid(row = 8, column = 0, columnspan = 4, sticky = "nsew")    
    chUpdateQNButton = tk.Button(left, text = "Get Chamber QNs", command = (lambda: updateQN(chamberPO, chamber, True, 0)))
    chUpdateQNButton.grid(row = 9, column = 0, sticky = "nsew")
    ctvUpdateQNButton = tk.Button(left, text = "Get CTV QNs", command = (lambda: updateQN(ctvPO, chamber, True, 1)))
    ctvUpdateQNButton.grid(row = 9, column = 1, sticky = "nsew")
    gpUpdateQNButton = tk.Button(left, text = "Get GasPanel QNs", command = (lambda: updateQN(gpPO, chamber, True, 2)))
    gpUpdateQNButton.grid(row = 9, column = 2, sticky = "nsew")
    allUpdateQNButton = tk.Button(left, text = "Get All QNs", command = (lambda: updateAllChQNs([chamberPO, ctvPO, gpPO], chamber)))
    allUpdateQNButton.grid(row = 9, column = 3, sticky = "nsew")
    viewQN(chamber, True,0)
    secondSeparator = ttk.Separator(left, orient = HORIZONTAL)
    secondSeparator.grid(row=11,column=0,columnspan=4,sticky="nsew" )
    checkRackStatusButton = tk.Button(left, text = "Check ERack for this system", command = (lambda: printStatus(system)))
    checkRackStatusButton.grid(row = 12, column = 0, columnspan = 4, sticky = "nsew")
    if chamberPO is not None:
        tlc_button = tk.Button(left, text = "Auto add TLC to chamber", command = (lambda: addtlc(chamberPO)))
        tlc_button.grid(row = 13, column = 0, columnspan=2, sticky = "nsew")
    manualTLCButton = tk.Button(left, text = "Manually add TLC to chamber", command = (lambda: opentlc(chamberPO)))
    manualTLCButton.grid(row = 13, column = 2, columnspan=2,sticky = "nsew")
    updateSystemNumEntry = tk.Entry(left)
    updateSystemNumEntry.grid(row = 14, column = 0,columnspan=2, sticky = "nsew")
    updateSystemNumButton = tk.Button(left, text = "Update System #", command = (lambda: setNewPO(chamber, updateSystemNumEntry, systemEntry, 0)))
    updateSystemNumButton.grid(row = 14, column = 2,columnspan=2, sticky = "nsew")
    updateChPOEntry = tk.Entry(left)
    updateChPOEntry.grid(row = 15, column = 0,columnspan=2, sticky = "nsew")
    updateChPOButton = tk.Button(left, text = "Update Chamber PO", command = (lambda: setNewPO(chamber, updateChPOEntry, POEntry, 1)))
    updateChPOButton.grid(row = 15, column = 2, columnspan = 2, sticky = "nsew")
    updateCTVPOEntry = tk.Entry(left)
    updateCTVPOEntry.grid(row = 16, column = 0,columnspan=2, sticky = "nsew")
    updateCTVPOButton = tk.Button(left, text = "Update CTV PO", command = (lambda: setNewPO(chamber, updateCTVPOEntry, ctvPOEntry, 2)))
    updateCTVPOButton.grid(row = 16, column = 2, columnspan = 2, sticky = "nsew")
    updateGPPOEntry = tk.Entry(left)
    updateGPPOEntry.grid(row = 17, column = 0, columnspan=2, sticky = "nsew")
    updateGPPOButton = tk.Button(left, text = "Update Gas Panel PO", command = (lambda: setNewPO(chamber, updateGPPOEntry, gpPOEntry, 3)))
    updateGPPOButton.grid(row = 17, column = 2, columnspan = 2, sticky = "nsew")
    createY7Button = tk.Button(left, text = "Create New QN",command = (lambda: createQN(chamberPO, chamber.system, "Y7")))
    createY7Button.grid(row = 18, column = 0,columnspan=4, sticky = "nsew")
    iomsButton = tk.Button(left, text = "Go to iOMS", command = (lambda: webbrowser.get(chrome).open_new_tab("http://ioms/MFG/ModuleStatus?PO=" + chamberPO + "#!/")))
    iomsButton.grid(row = 19, column = 0, columnspan = 4, sticky = "nsew")
    create3DButton = tk.Button(left, text = "Create new 3D form", command = (lambda: webbrowser.get(chrome).open_new_tab("http://sppartner/sites/Global3D/Lists/VMORevision/Item/newifs.aspx")))
    create3DButton.grid(row = 20, column = 0, columnspan=4, sticky = "nsew")
    copyQNsButton = tk.Button(left, text = "Copy QNs to clipboard", command = (lambda: copyQNsToClipboard(chamber)))
    copyQNsButton.grid(row = 21, column = 0, columnspan = 4, sticky = "nsew")


#Displays QNs in notebook 
def viewQN(ch, onlyOpen, qnId):
    status = ""
    qnIdx = 0
    if onlyOpen:
        status = "Open QNs for " + ch.system + \
                 "\n############################################################"
    else:
        status = "All QN's for " + ch.system
    n = ttk.Notebook(left, height= 388)
    f1 = ttk.Frame(n)
    f2 = ttk.Frame(n)
    f3 = ttk.Frame(n)
    f4 = ttk.Frame(n)
    f5 = ttk.Frame(n)
    f6 = ttk.Frame(n)
    n.add(f1, text = " Chamber QNs ")
    n.add(f2, text = " CTV QNs ")
    n.add(f3, text = " Gas Panel QNs ")
    n.add(f4, text = " Chamber ILs ")
    n.add(f5, text = " CTV ILs ")
    n.add(f6, text = " Gas Panel ILs ")
    n.grid(row = 10, column = 0, columnspan=4,sticky="nsew")
    frames = [f1, f2, f3, f4, f5, f6]
    for qnIdx in range(3):
        openQNs = ch.openQNs(qnIdx)
        openInspLots = ch.openInspLots(qnIdx)
        if openQNs:
            n.tab(qnIdx, text = (n.tab(qnIdx)['text'] + "(" + str(len(openQNs))) + ")")
        if openInspLots:
            n.tab(qnIdx + 3, text = (n.tab(qnIdx + 3)['text'] + "(" + str(len(openInspLots))) + ")")
        numOpen = len(openQNs)
        numOpenILs = len(openInspLots)
        if numOpen != 0:
            qnLinks = []
            texts = []
            for i in range(numOpen):
                separator = ttk.Separator(frames[qnIdx], orient = HORIZONTAL)
                separator.pack(side='top', fill='both', expand=True)
                texts.append(tk.Text(frames[qnIdx], height = 2, borderwidth = 0))
                qnLinks.append(texts[i])
                qnLinks[i].tag_configure("center", justify='center')
                qnLinks[i].tag_add("center", "1.0", "end")
                qnLinks[i].configure(bg=root.cget('bg'), relief="flat")
                qnLinks[i].pack()
                hyperlink = HyperlinkManager(qnLinks[i])
                qnLinks[i].insert(INSERT, openQNs[i].Type + " ")
                qnLinks[i].insert(INSERT, openQNs[i].QNNum, hyperlink.add(lambda x = openQNs[i].QNNum: openQN(x)))
                qnLinks[i].insert(INSERT, " - " + openQNs[i].Desc.upper())
                qnLinks[i].configure(state="disabled")
                ltsText = tk.Text(frames[qnIdx], height = 2, borderwidth =0)
                ltsText.tag_configure("center", justify='center')
                ltsText.tag_add("center", "1.0", "end")
                ltsText.configure(bg=root.cget('bg'), relief="flat")
                #need to disable lts text
                ltsText.pack()
                if openQNs[i].lastScanned:
                    ltsText.insert(INSERT, openQNs[i].lastScanned)
                if openQNs[i].partNum and (openQNs[i].Type == "Y8" or openQNs[i].Type == "YI"):
                    ltsButton = tk.Button(frames[qnIdx], text="Refresh last scanned", command = (lambda q = openQNs[i], text = ltsText: getLastScanned(ch.chPO, q, text)))
                    ltsButton.pack()
        else:
            tk.Label(frames[qnIdx], text = "No open QN data").pack()
        if numOpenILs != 0:
            n.select(qnIdx + 3)
            ilLinks = []
            ilTexts = []
            for i in range(numOpenILs):
                separator = ttk.Separator(frames[qnIdx + 3], orient = HORIZONTAL)
                separator.pack(side='top', fill='both', expand=True)
                ilTexts.append(tk.Text(frames[qnIdx + 3], height = 2, borderwidth = 0))
                ilLinks.append(ilTexts[i])
                ilLinks[i].tag_configure("center", justify='center')
                ilLinks[i].tag_add("center", "1.0", "end")
                ilLinks[i].configure(bg=root.cget('bg'), relief="flat")
                ilLinks[i].pack()
                ilhyperlink = HyperlinkManager(ilLinks[i])
                ilLinks[i].insert(INSERT, openInspLots[i].lotNum, ilhyperlink.add(lambda x = openInspLots[i].lotNum: openInspLot(x)))
                ilLinks[i].insert(INSERT, " - " + openInspLots[i].Desc.upper())
                ilLinks[i].configure(state="disabled")
        else:
            tk.Label(frames[qnIdx + 3], text = "No open InspLot data").pack()
    n.select(qnId)


##################################################################################
#                                                                                #
#                               Chrome Dependent                                 #
#                                                                                #
##################################################################################
def updateAllDataForAllPorts(bay, sectionTexts, statusTexts, updateTimeText, percentTexts):
    global floorDataUpdated
    driver = createDriver(True)
    start = time.time()
    for b in bays[1:]:
        updateAllQNsForAllPorts(b.bayNumber, driver)
    print("QNS DONE AT " + datetime.now().strftime("%m/%d/%Y %H:%M:%S"))
    end = time.time()
    print("Took " + str(end-start) + " seconds")
    for b in bays[1:]:
        updateLTSForAllOpenQNs(b.bayNumber, driver)
    print("LTS DONE AT " + datetime.now().strftime("%m/%d/%Y %H:%M:%S"))
    for b in bays[1:]:
        updateMayanStatusForAllPorts(b.bayNumber, driver)
        getBuildStatus(b.bayNumber, driver)
    print("MAYAN/BUILD DONE AT " + datetime.now().strftime("%m/%d/%Y %H:%M:%S"))
    end = time.time()
    print("Took " + str(end-start) + " seconds")
    drawBuildStatus(bay, percentTexts)
    drawMayanStatus(bay, sectionTexts, statusTexts, updateTimeText)
    driver.quit()
    floorDataUpdated = [True, datetime.now().strftime("%m/%d/%Y %I:%M%p")]


def updateAllQNsForAllPorts(bay_num, driver):
    for ch in bays[bay_num].chambers:
        #print(ch.system)
        updateAllChQNs2(driver, [ch.chPO, ch.ctvPO, ch.gpPO], ch)

def updateAllDataForBay(bay, sectionTexts, statusTexts, updateTimeText, percentTexts):
    global bayDataUpdated
    driver = createDriver(True)
    for ch in bays[bay].chambers:
        updateAllChQNs2(driver, [ch.chPO, ch.ctvPO, ch.gpPO], ch)
    print("QNs done")
    updateLTSForAllOpenQNs(bay, driver)
    print("LTS done")
    updateMayanStatusForAllPorts(bay, driver)
    print("Mayan done")
    getBuildStatus(bay, driver)
    print("Build done")
    driver.quit()
    drawBuildStatus(bay, percentTexts)
    drawMayanStatus(bay, sectionTexts, statusTexts, updateTimeText)
    bayDataUpdated = [True, datetime.now().strftime("%m/%d/%Y %I:%M%p")]


def updateAllQNsOnAllChambers(bay_num):
    driver = createDriver(True)
    for ch in bays[bay_num].chambers:
        print(ch.system)
        updateAllChQNs2(driver, [ch.chPO, ch.ctvPO, ch.gpPO], ch)
    driver.quit()


##Merge these two
def updateAllChQNs(pos, chamber):
    driver = createDriver(True)
    for i in range(3):
        if pos[i] and (pos[i] != "XXXXX"):
            updateQN(pos[i], chamber, True, i, driver)
    driver.quit()


#Same as updateAllQNs but called when a driver already exists
def updateAllChQNs2(driver, pos, chamber):
    for i in range(3):
        if pos[i] and (pos[i] != "XXXXX"):
            updateQN(pos[i], chamber, True, i, driver)


#Gets QNs and InspLots for a PO
def updateQN(po, ch, viewQNs, qnIdx, driver = None):
    if po and (po != "XXXXX"):
        print("Checking QNs for PO: " + po + "...")
        externalDriver = True
        qnCount = 0
        if not driver:
            options = Options()
            options.headless = True # set to False to see chrome window while running
            options.add_argument("--window-size=1920,1200")
            DRIVER_PATH = r"./driver/chromedriver.exe"
            driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
            externalDriver = False
        driver.get("http://webmprd:3333/rest/AMAT_iOMS_SSG/api/getQN_LOT_DataFromSAP?ORD_NUMBER=" + po)
        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))#
            x = driver.find_element_by_css_selector("body > pre").text#
            y = json.loads(x)#
            ch.allQNs[qnIdx].clear()
            if y["SAP_QN_DATA_RESPONSE"]["RESPONSE"]["QL_COUNT"] != "0":#
                for q in y["SAP_QN_DATA_RESPONSE"]["RESPONSE"]["QN_DATA"]["QLNOTE"]:#
                    posPartNums = []
                    status = True if (q["ZZIMMFIX"] != "true") else False#
                    partNum = q["MATNR"]
                    if qnIdx == 0:
                        if status:
                            #print(partNum)
                            PN = fnmatch.filter(partNum.split(), '????-?????')
                            if PN:
                                partNum = PN
                                ch.QNs.append(QN(q["QMNUM"], q["QMART"], q["QMTXT"], status, partNum[0]))#
                                continue
                            else:
                                PNReg = re.compile(r'\d{4}-\d{5}')
                                foundPN = PNReg.search(q["QMTXT"])#
                                if foundPN:
                                    #print(q["QMNUM"])#
                                    partNum = foundPN.group()
                                    #print(partNum)
                                    ch.QNs.append(QN(q["QMNUM"], q["QMART"], q["QMTXT"], status, partNum))#
                                    continue
                        ch.QNs.append(QN(q["QMNUM"], q["QMART"], q["QMTXT"], status))#
                        qnCount = qnCount + 1
                    else:
                        ch.allQNs[qnIdx].append(QN(q["QMNUM"], q["QMART"], q["QMTXT"], status))           
            ch.allInspLots[qnIdx].clear()
            if y["SAP_QN_DATA_RESPONSE"]["RESPONSE"]["INSLOT_COUNT"] != "0":
                for i in y["SAP_QN_DATA_RESPONSE"]["RESPONSE"]["INSPECTION_DATA"]["INSLOT"]:
                    status = True if (i["VBEWERTUNG"] != "A") else False
                    ch.allInspLots[qnIdx].append(InspLot(i["PRUEFLOS"], i["KTEXT"], i["VBEWERTUNG"], status))             
        except:
            ch.allQNs[qnIdx].clear()
            ch.allInspLots[qnIdx].clear()
            if not externalDriver:
                driver.quit()
        if not externalDriver:
                driver.quit()
        if viewQNs:
            viewQN(ch, True, qnIdx)


def updateLTSForAllOpenQNs(bay, driver):
    for chamber in bays[bay].chambers:
        for qn in chamber.openQNs(0):
            if qn.partNum and (qn.Type == "YI" or qn.Type == "Y8"):
                updateLastScanned(chamber.chPO, qn, driver)


def updateLastScanned(po, qn, driver):
    print("Checking LTS for " + qn.partNum + "...")
    link = "http://dca-app-833/LTSWeb/api/LTSPACKAGEAPI/SearchPackageDetails?RFID=&PartNumber=" + str(qn.partNum) + "&StockroomReq=&ProdOrder=" + po + "&WorkCenter=&Operation=&JobNumber="
    driver.get(link)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))#
    x = driver.find_element_by_css_selector("body > pre").text#
    y = json.loads(x)
    if y["recordsTotal"] == 0:
        qn.lastScanned = "No Data Found"
    else:
        try:
            qn.lastUpdateTimestamp = datetime.now().strftime("%m/%d/%Y %I:%M%p")
            loc = y["ltsPackagedata"][0]["LTS_LOCATIONDESIGNATION"]
            lastSeen = formatLTSDateAndTime(y["ltsPackagedata"][0]["LAST_SEEN_TIME"])#[0:10]
            qn.lastScanned = (str(qn.partNum) + " Last seen " + lastSeen + " " + loc + "\nChecked LTS on: " + qn.lastUpdateTimestamp)
        except:
            qn.lastScanned = "No Data Found"
    #print(qn.lastScanned)
    return

def formatLTSDateAndTime(dateTimeString):
    formattedDateTime = datetime.strptime(dateTimeString[0:18].replace("T", " "), '%Y-%m-%d %H:%M:%S')
    localFormattedDateTime = formattedDateTime.replace(tzinfo=timezone.utc).astimezone(tz=None)
    strFormattedDateTime = localFormattedDateTime.strftime("%m/%d/%Y %I:%M%p")
    return strFormattedDateTime
    

#Query LTS to find th last known location of a part that belongs to a specific PO
def getLastScanned(po, qn, text):
    print("Checking LTS for " + qn.partNum + "...")
    options = Options()
    options.headless = True # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    link = "http://dca-app-833/LTSWeb/api/LTSPACKAGEAPI/SearchPackageDetails?RFID=&PartNumber=" + str(qn.partNum) + "&StockroomReq=&ProdOrder=" + po + "&WorkCenter=&Operation=&JobNumber="
    driver.get(link)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))#
    x = driver.find_element_by_css_selector("body > pre").text#
    y = json.loads(x)
    if y["recordsTotal"] == 0:
        driver.quit()
        text.delete('1.0', END)
        text.insert(INSERT, "No Data Found")
        qn.lastScanned = "No Data Found"
    else:
        try:
            qn.lastUpdateTimestamp = datetime.now().strftime("%m/%d/%Y %I:%M%p")
            loc = y["ltsPackagedata"][0]["LTS_LOCATIONDESIGNATION"]
            lastSeen = formatLTSDateAndTime(y["ltsPackagedata"][0]["LAST_SEEN_TIME"])#[0:10]
            text.delete('1.0', END)
            qn.lastScanned = (str(qn.partNum) + " Last seen " + lastSeen + " " + loc + "\nChecked LTS on: " + qn.lastUpdateTimestamp)
            text.insert(INSERT, qn.lastScanned)
            driver.quit()
        except:
            qn.lastScanned = "No Data Found"
            text.delete('1.0', END)
            text.insert(INSERT, "No Data Found")
            driver.quit()
    return



def getBuildStatus(bay, driver):
    for chamber in bays[bay].chambers:
        chamberPO = chamber.chPO
        try:
            driver.get("http://eagleeye_api/EagleEye/api/v1//OperationBuild/" + chamberPO)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))
            x = driver.find_element_by_css_selector("body > pre").text
            y = json.loads(x)
            total = 0
            count = 0
            for OP in y:
                if OP["POB_ID"] != 0 and OP["Operation"] not in ("Module Test", "CP", "F50"):
                    total = total + OP["Build_Percentage"]
                    count = count + 1
            overallPercent = int(total/count)
            chamber.buildPercent = str(overallPercent)
        except:
            pass


#Get estimate of build percent
def updateBuildStatus(bay_num, canvas, percentTexts):
    driver = createDriver(True)
    for chamber in bays[bay_num].chambers:
        chamberPO = chamber.chPO
        try:
            driver.get("http://eagleeye_api/EagleEye/api/v1//OperationBuild/" + chamberPO)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))
            x = driver.find_element_by_css_selector("body > pre").text
            y = json.loads(x)
            total = 0
            count = 0
            for OP in y:
                if OP["POB_ID"] != 0 and OP["Operation"] not in ("Module Test", "CP", "F50"):
                    total = total + OP["Build_Percentage"]
                    count = count + 1
            overallPercent = int(total/count)
            chamber.buildPercent = str(overallPercent)
        except:
            pass
    drawBuildStatus(bay_num, percentTexts)
    driver.quit()


def updateMayanStatusForAllPorts(bay_num, driver):
    for chamber in bays[bay_num].chambers:
        chamberPO = chamber.chPO
        try:
            driver.get("http://eagleeye_api/EagleEye/api/v1//MayanDetails?Production_Order=" + chamberPO)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))
            x = driver.find_element_by_css_selector("body > pre").text
            y = json.loads(x)
            section = y[0]["Current_Mayan_Section"].replace("-", " ")
            chamber.currentMayanSection = section
            currentTest = y[0]["Current_Mayan_Test"]
            chamber.currentMayanTest = currentTest
            chamber.mayanUpdateTimestamp = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        except:
            pass


#Get current Mayan test locations for all chambers in a bay
def updateMayanStatus(bay_num, canvas, sectionTexts, statusTexts, updateTimeText):
    driver = createDriver(True)
    for chamber in bays[bay_num].chambers:
        chamberPO = chamber.chPO
        try:
            driver.get("http://eagleeye_api/EagleEye/api/v1//MayanDetails?Production_Order=" + chamberPO)
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > pre")))
            x = driver.find_element_by_css_selector("body > pre").text#
            y = json.loads(x)
            section = y[0]["Current_Mayan_Section"].replace("-", " ")
            chamber.currentMayanSection = section
            currentTest = y[0]["Current_Mayan_Test"]
            chamber.currentMayanTest = currentTest
            chamber.mayanUpdateTimestamp = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
            drawMayanStatus(bay_num, sectionTexts, statusTexts, updateTimeText)
        except:
            pass
    canvas.itemconfig(updateTimeText, text="Mayan updated " + datetime.now().strftime("%m/%d/%Y %H:%M:%S"))
    driver.quit()


#Get current TLC hours logged
def getUserTLC(driver):
    global today
    global empNum
    driver.get("http://dca-app-1445/TLCAPI/api/TLCLogHoursApi/?userid=" + empNum + "&currentDate=" + today + "&plant=null")
    x = driver.find_element_by_css_selector("body > pre").text
    y = json.loads(x)
    hrs = 0
    for _ in y:
        hrs = hrs + float(_["sapDecimalHours"])
    print(hrs)
    return int(hrs)


#Open TLC for manual adding
def opentlc(po):
    r = Tk()
    r.withdraw()
    r.clipboard_clear()
    r.clipboard_append(po)
    r.update()
    r.destroy()
    webbrowser.get(chrome).open_new_tab("http://dca-wb-281/TLC/TLCLogHours/LogHoursSummary")


#Auto add TLC to keep user in "healthy range"
def addtlc(po):
    options = Options()
    options.headless = False # set to False to see chrome window while running
    options.add_argument("--window-size=1920,1200")
    DRIVER_PATH = r"./driver/chromedriver.exe"
    driver = webdriver.Chrome(options=options, executable_path=resource_path(DRIVER_PATH))
    hrs = getUserTLC(driver)
    print(hrs)
    mins = "%02d" % random.randint(0,30)
    print(mins)
    driver.get("http://dca-wb-281/TLC/TLCLogHours/LogHoursSummary")
    print(hrs)
    if hrs < 8:
        hrsneeded = 8 - hrs
        print(hrsneeded)
        try:
            addButton = WebDriverWait(driver, 35).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#AddHoursid')))
            addButton.click()
            time.sleep(1)
            tlcPO = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#tlc_prd_orders')))
            tlcPO.click()
            time.sleep(1)
            tlcPO.send_keys(po)
            tlcPOSelect = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#ui-id-7 > li')))
            print("po loaded")
            tlcPhase = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#Phase')))
            tlcPhase.click()
            phaseSelect = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#Phase > option:nth-child(4)")))
            phaseSelect.click()
            time.sleep(1)
            tlcOP = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#OperationId')))
            tlcOP.click()
            ops = tlcOP.find_elements_by_tag_name("option")
            opSelect = ops[1]
            opSelect.click()
            time.sleep(1)
            tlcInstruction = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#OmsIndexId')))
            tlcInstruction.click()
            instructionSelect = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="200"]')))
            instructionSelect.click()
            startTime = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#timepicker_start')))
            startTime.click()
            startTime.send_keys("0600")
            endTime = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#timepicker_end')))
            endTime.click()
            if hrsneeded >= 4:
                endTime.send_keys(str(hrsneeded + 6) + mins)
            else:
                endTime.send_keys("0" + str(hrsneeded + 6) + mins)
##            saveButton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#btnSave')))
##            saveButton.click()
##            time.sleep(2)
##            WebDriverWait(driver, 35).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#AddHoursid')))
##            print("successfully added " + hrs + mins)
        except:
            print("Error adding TLC, please add manually.")


def createQN(po, system, qnType):
    global user
    x = system.find("-")
    system = system[0:x]
    webbrowser.get(chrome).open_new_tab("http://qualitynotescm/QN?productionOrder=" + po +"&ntLoginID=" + user + "&application=prompt&slot=" + system + "&plant=4070&material=")


def openQN(qn):
    webbrowser.get(chrome).open_new_tab("https://epvpwd.amat.com:8065/com.amat.irj.portal?app=ChgQaNotif?RIWO00-QMNUM=0000" + qn)


def openInspLot(il):
    webbrowser.get(chrome).open_new_tab("https://epvpwd.amat.com:8065/irj/portal/DisIL?QALS-PRUEFLOS=" + il)


##################################################################################
#                                                                                #
#                                     Admin                                      #
#                                                                                #
##################################################################################
def generateFloorReport():
    global floorDataUpdated
    ports = ["A", "B", "C", "D", "E", "F", "G", "H"]
    if floorDataUpdated[0]:
        with open("Report.txt", "w") as report:
            result = "Data Updated " + floorDataUpdated[1] + "\n"
            for b in bays[1:]:
                result = result + "-----------------------------------------------------------\n" + "Bay " + str(b.bayNumber) + "\n-----------------------------------------------------------\n"
                for ch in b.chambers:
                    if ch.system != "XXXXX-X":
                        closedQNs = ""
                        result = result + ch.port + ":\n" + ch.system + " (" + ch.chType + ") " + "PO: " + ch.chPO + " \n"
                        data = df.loc[df['System #'].astype(str).str.contains(ch.system)]
                        try:
                            chStatus = data.values[0][2]
                        except:
                            chStatus = "ERROR"
                            pass
                        result = result + "Passdown: " + chStatus + "\nMayan Status: " + ch.currentMayanSection + "\n\n   QNs:\n"
                        for qn in ch.QNs:
                            if qn.isOpen:
                                result = result + "\t" + qn.print()
                                if qn.lastScanned:
                                    result = result + "\t-->\t" + qn.lastScanned.replace("\n", " | ") + "\n"
                            elif qn.Type == "YI" or qn.Type == "Y8":
                                closedQNs = closedQNs + "\t" + qn.print()
                        result = result + "\n" + closedQNs + "\n"
                        for il in ch.openInspLots(0):
                            result = result + "\t" + il.print()
                        result = result + "\n"
                    else:
                        result = result + ch.port + ": " + "Empty\n\n"
            report.write(result)
        print("Floor report generated")
    else:
        print("Please udpate all data")


def generateBayReport(bay):
    global bayDataUpdated
    ports = ["A", "B", "C", "D", "E", "F", "G", "H"]
    if bayDataUpdated[0]:
        print(bayDataUpdated[1])
        with open("Bay" + str(bay) + "Report.txt", "w") as bayReport:
            result = "Data Updated " + bayDataUpdated[1] + "\n"
            result = result + "-----------------------------------------------------------\n" + "Bay " + str(bay) + "\n-----------------------------------------------------------\n"
            for ch in bays[bay].chambers:
                if ch.system != "XXXXX-X":
                        closedQNs = ""
                        result = result + ch.port + ":\n" + ch.system + " (" + ch.chType + ") " + "PO: " + ch.chPO + " \n"
                        data = df.loc[df['System #'].astype(str).str.contains(ch.system)]
                        try:
                            chStatus = data.values[0][2]
                        except:
                            chStatus = "ERROR"
                            pass
                        result = result + "Passdown: " + chStatus + "\nMayan Status: " + ch.currentMayanSection + "\n\n   QNs:\n"
                        for qn in ch.QNs:
                            if qn.isOpen:
                                result = result + "\t" + qn.print()
                                if qn.lastScanned:
                                    result = result + "\t-->\t" + qn.lastScanned.replace("\n", " | ") + "\n"
                            elif qn.Type == "YI" or qn.Type == "Y8":
                                closedQNs = closedQNs + "\t" + qn.print()
                        result = result + "\n" + closedQNs + "\n"
                        for il in ch.openInspLots(0):
                            result = result + "\t" + il.print()
                        result = result + "\n"
                else:
                    result = result + ch.port + ": " + "Empty\n\n"
            bayReport.write(result)
            print("Bay report generated")
    else:
        print("Please update all bay data")



############################################################ MAIN PROGRAM BEGINS ########################################################################
#pd.set_option('display.max_columns', 30)
manualPOsFile = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOsUnited.xlsx'
passdownPath = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\SharepointTest\LEADS PASSDOWN*.xlsx'
crossoverPath = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\(DTF Checklists - TEST QUEUE)\(TEST QUEUE )\TEST QUEUE*.xlsx'
manualPOPath = (r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOsUnited.xlsx')
chrome = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"
chamber_image_file = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\chamber.png'
bay_image_file = r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\baydrawingbagel.png'
############################For home dev uses local verison of passdown and test queue update required daily#############################################
##passdownPath = r'C:\Users\icarrillo170821\Documents\home\LEADS PASSDOWN*.xlsx'
##crossoverPath = r'C:\Users\icarrillo170821\Documents\home\TEST QUEUE*.xlsx'
#########################################################################################################################################################
config = configparser.ConfigParser()
config.read('config.ini')
d = dict(config.items('GENERAL'))
for key in d:
    d[key] = d[key].strip("|")
#print(d)
bay_num_str = d['bay']
bay_num = int(d['bay'])
#General Purpose Global
##bay_num_str = "8"
##bay_num = 8
chambers = []
cells = []
active_buttons = []
chamber_locations = [1,1,1,1,1,1,1,1]
ports = ["A","B","C","D","E","F", "G", "H"]
bays = ["dummy", Bay(1,6), Bay(2,6), Bay(3,6), Bay(4,6), Bay(5,6), Bay(6,8), Bay(7,6), Bay(8,6), Bay(9,8), Bay(10,8)]
today = datetime.today().strftime("%m/%d/%y")
print(today)
floorDataUpdated = [False, ""]
bayDataUpdated = [False, ""]
skip = True
isContractor = False
user = os.getlogin()
if fnmatch.fnmatch(user, "*x[0-9]*"):
    isContractor = True
    empNum = "x" + re.sub("[^0-9]", "", user)
else:
    empNum = re.sub("[^0-9]", "", user)

#Get ERack Status
rackStatus = load_workbook(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\EQRK Status.xlsx')
rackStatusSheet = rackStatus["Sheet1"]

#Find latest crossover/passdown files and update path
crossoverFile = getLatestFile(crossoverPath)
print(crossoverFile)
crossoverPath = crossoverPath.replace("TEST QUEUE*.xlsx", crossoverFile)
passdownFile = getLatestFile(passdownPath)
print(passdownFile)
passdownPath = passdownPath.replace("LEADS PASSDOWN*.xlsx", passdownFile)

#Create a datarame with columns System#, Chamber PO#, CH Type
#Drop extra lines at beginning of excel sheet
data = pd.read_excel(crossoverPath, sheet_name= 'QUEUE', usecols = 'G:I', dtype=str)
dfCrossover = pd.DataFrame(data)
startIndex = dfCrossover.loc[dfCrossover['Unnamed: 6'] == 'Slot /Sys - Ch# '].index[0]
dfCrossover.drop(dfCrossover.index[:startIndex+1], inplace=True)
dfCrossover.reset_index(drop=True,inplace=True)
dfCrossover.rename(columns={"Unnamed: 6": "Slot /Sys - Ch# ", "Unnamed: 7": "Build PO#", "Unnamed: 8":"CH Type"}, inplace=True)

#Create dataframe for leads passdown
data = pd.read_excel(passdownPath)
df = pd.DataFrame(data, columns= ['Bay ','System #', 'Status of Chamber', 'Issues', 'START Date','Port Days'])
df.fillna('', inplace=True)


#Create dataframe for manual po workbook
data = pd.read_excel(manualPOPath, sheet_name = 'Sheet1', usecols = 'A:E', dtype = str)
manualDF = pd.DataFrame(data)
#Remove entries that don't match leads passdown
manualPOs = load_workbook(manualPOsFile)
ws = manualPOs.active
rows = list(ws.rows)
for i in range(66):
    if df.at[i, "System #"].strip() != manualDF.at[i, "System"]:
        for x in range(2,6):
            ws.cell(column=x, row=i + 2, value="")
manualPOs.save(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\ManualPOs.xlsx')
    
getPOs(chambers)
getPriorityColors(passdownPath, chambers)
for c in chambers:
    bays[c.bay].chambers.append(c)
chamber_locations = findChamberLocations(bay_num)
print(chamber_locations)
root = tk.Tk()
root.title("Bagel - Beta")
try:
    root.iconbitmap(r'\\amat.com\Folders\Austin\Global-Ops\AMO\CPI_TestWorkCntr\TECH FOLDERS\ Irvin Carrillo\Bagel.ico')
except:
    pass
w = 1654
h = 1000
ws = root.winfo_screenwidth() # width of the screen
hs = root.winfo_screenheight() # height of the screen
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
root.geometry('%dx%d+%d+%d' % (w, h, x, y-40))
root.resizable(False, False)
bay_image = ImageTk.PhotoImage(file = bay_image_file)


topframe = tk.Frame(root, height=40, bg="#4599C3", bd=1, relief=SOLID)
topframe.columnconfigure(7, weight=1)
middleframe = tk.PanedWindow(root, height =950)#845
topframe.pack(side="top", fill="x", expand=True)
middleframe.pack(fill="both")#side="top", fill="both", expand=True)
left = tk.Frame(middleframe, width=500,height=950, bg="blue", bd=1, relief=SOLID)
left.columnconfigure(3,weight=1)
#left.columnconfigure(1,weight=1)
middleframe.paneconfig(left, minsize=500)
right = tk.Frame(middleframe, width =1154,height=950, bg="red",bd=1,relief=SOLID)
middleframe.paneconfig(right, minsize=1154)
middleframe.add(left)
middleframe.add(right)
canvas = Canvas(right, width=1154, height=950)
canvas.create_image(575,470, image = bay_image)
mayanSectionTexts = []
mayanStatusTexts = []
buildPercentTexts = []
buildLabelTexts = []
aMayanSection = canvas.create_text(680,40, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
bMayanSection = canvas.create_text(680,290, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
cMayanSection = canvas.create_text(680,520, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
dMayanSection = canvas.create_text(680,750, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
eMayanSection = canvas.create_text(450,750, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
fMayanSection = canvas.create_text(450,520, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
gMayanSection = canvas.create_text(450,280, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
hMayanSection = canvas.create_text(450,40, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="")
mayanSectionTexts.extend([aMayanSection,bMayanSection,cMayanSection,dMayanSection,eMayanSection,fMayanSection,gMayanSection,hMayanSection])
aMayanStatus = canvas.create_text(680,110, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
bMayanStatus = canvas.create_text(680,360, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
cMayanStatus = canvas.create_text(680,590, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
dMayanStatus = canvas.create_text(680,820, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
eMayanStatus = canvas.create_text(450,820, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
fMayanStatus = canvas.create_text(450,590, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
gMayanStatus = canvas.create_text(450,350, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
hMayanStatus = canvas.create_text(450,110, width =205, justify = tk.CENTER, fill="black", font="Helvetica 12", text="")
mayanStatusTexts.extend([aMayanStatus,bMayanStatus,cMayanStatus,dMayanStatus,eMayanStatus,fMayanStatus,gMayanStatus,hMayanStatus])
mayanUpdateTimeText = canvas.create_text(555,920, justify=tk.CENTER, fill="black", font = "Helvetica 12 bold", text ="Mayan progress not updated")
aBuildPercent = canvas.create_text(1060,100, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
bBuildPercent = canvas.create_text(1060,350, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
cBuildPercent = canvas.create_text(1060,580, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
dBuildPercent = canvas.create_text(1060,810, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
eBuildPercent = canvas.create_text(75,810, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
fBuildPercent = canvas.create_text(75,580, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
gBuildPercent = canvas.create_text(75,350, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
hBuildPercent = canvas.create_text(75,110, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="-%")
buildPercentTexts.extend([aBuildPercent,bBuildPercent,cBuildPercent,dBuildPercent,eBuildPercent,fBuildPercent,gBuildPercent,hBuildPercent])
aBuildLabel = canvas.create_text(1060,70, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
bBuildLabel = canvas.create_text(1060,320, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
cBuildLabel = canvas.create_text(1060,550, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
dBuildLabel = canvas.create_text(1060,780, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
eBuildLabel = canvas.create_text(75,780, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
fBuildLabel = canvas.create_text(75,550, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
gBuildLabel = canvas.create_text(75,320, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
hBuildLabel = canvas.create_text(75,70, width =205, justify = tk.CENTER, fill="black", font="Helvetica 14 bold", text="Build:")
buildPercentTexts.extend([aBuildLabel,bBuildLabel,cBuildLabel,dBuildLabel,eBuildLabel,fBuildLabel,gBuildLabel,hBuildLabel])

canvas.pack()

currentBayLabel = tk.Label(topframe, bg="#4599C3",font="Helvetica 32 bold", text = "Now Viewing Bay " + bay_num_str)
currentBayLabel.grid(row = 0, column = 0, rowspan=2, sticky = "w")
new_bay_entry = tk.Entry(topframe)
new_bay_entry.grid(row = 0, column = 2, sticky = "nsew")
refresh_button = tk.Button(topframe, text = "Change bay", command = (lambda: change_bay(root, chamber_image, new_bay_entry.get(), new_bay_entry, active_buttons,\
                                                                                       currentBayLabel, canvas, mayanSectionTexts, mayanStatusTexts, mayanUpdateTimeText, buildPercentTexts)))#locations
refresh_button.grid(row = 0, column = 3, sticky = "nsew")
updateAllDataButton = tk.Button(topframe, text = "Update all data for all ports", command = (lambda: updateAllDataForAllPorts(bay_num, mayanSectionTexts, mayanStatusTexts, mayanUpdateTimeText, buildPercentTexts)))
updateAllDataButton.grid(row = 0, column = 4, sticky = "nsew")
generateFloorReportButton = tk.Button(topframe, text = "Generate Floor Report", command = (lambda: generateFloorReport()))
generateFloorReportButton.grid(row = 0, column = 5, sticky = "nsew")
updateAllDataForBayButton = tk.Button(topframe, text = "Update all data for bay", command = (lambda: updateAllDataForBay(bay_num, mayanSectionTexts, mayanStatusTexts, mayanUpdateTimeText, buildPercentTexts)))
updateAllDataForBayButton.grid(row = 0, column = 6, sticky = "nsew")
generateBayReportButton = tk.Button(topframe, text = "Generate Bay Report", command = (lambda: generateBayReport(bay_num)))
generateBayReportButton.grid(row = 0, column = 7, sticky = "nsew")
rackStatusEntry = tk.Entry(topframe)
rackStatusEntry.grid(row = 1, column = 2, sticky = "nsew")
rackStatusButton = tk.Button(topframe, text = "Check ERack status for system", command = (lambda: printStatus(rackStatusEntry.get().upper())))
rackStatusButton.grid(row = 1, column = 3, sticky = "nsew")
rackRefreshButton = tk.Button(topframe, text = "Refresh ERack status", command = (lambda: refreshRackStatus()))
rackRefreshButton.grid(row = 1, column = 4, sticky = "nsew")
##updateAllBayQNsButton = tk.Button(topframe, text = "Update all Chamber QNs", command = (lambda: updateAllBayQNs(bay_num)))
##updateAllBayQNsButton.grid(row = 1, column = 5,sticky="nsew")
updateAllBayQNsButton = tk.Button(topframe, text = "Update all QNs for bay", command = (lambda: updateAllQNsOnAllChambers(bay_num)))
updateAllBayQNsButton.grid(row = 1, column = 5,sticky="nsew")
updateMayanBayButton = tk.Button(topframe, text = "Update Mayan Progress", command = (lambda: updateMayanStatus(bay_num, canvas, mayanSectionTexts, mayanStatusTexts, mayanUpdateTimeText)))
updateMayanBayButton.grid(row = 1, column = 6, sticky = "nsew")
updateBuildButton = tk.Button(topframe, text = "Update Build %", command = (lambda: updateBuildStatus(bay_num, canvas, buildPercentTexts)))
updateBuildButton.grid(row = 1, column = 7, sticky = "nsew")
userLabel = tk.Label(topframe, bg="#4599C3",font="Helvetica 16", text = "Signed in as: " + user)
userLabel.grid(row = 1, column = 8, sticky = "e")
chamber_image = PhotoImage(file = chamber_image_file)
create_buttons(root, chamber_image, chamber_locations, active_buttons)
count = 0
while(True):
    try:
        create_window_Generic(bay_num, count)
        break
    except:
        count = count + 1

root.mainloop()
